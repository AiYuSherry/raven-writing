"""Input reader — detects and reads various input types including directories."""

import csv
import os
import re
import zipfile
from html import unescape


URL_PATTERNS = [
    (re.compile(r'^https?://mp\.weixin\.qq\.com/'), 'weixin'),
    (re.compile(r'^https?://zhuanlan\.zhihu\.com/'), 'zhihu'),
    (re.compile(r'^https?://'), 'url'),
]

# File extensions to read when scanning a directory
READABLE_EXTENSIONS = {
    '.txt', '.md', '.py', '.json', '.yml', '.yaml', '.toml', '.cfg', '.ini',
    '.html', '.css', '.js', '.jsx', '.ts', '.tsx', '.vue', '.svelte',
    '.java', '.kt', '.go', '.rs', '.rb', '.php', '.swift',
    '.c', '.h', '.cpp', '.hpp',
    '.xml', '.svg', '.yaml', '.sh', '.bash', '.zsh',
    '.conf', '.env.example', '.gitignore', '.dockerfile',
    '.csv', '.sql', '.r', '.lua', '.epub',
}

SPREADSHEET_EXTENSIONS = {'.xlsx', '.xlsm', '.xls'}


def _strip_html_text(html):
    text = re.sub(r"(?is)<script.*?>.*?</script>", " ", html)
    text = re.sub(r"(?is)<style.*?>.*?</style>", " ", text)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</p\s*>", "\n\n", text)
    text = re.sub(r"(?i)</div\s*>", "\n", text)
    text = re.sub(r"(?i)</h[1-6]\s*>", "\n\n", text)
    text = re.sub(r"(?is)<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def read_epub(path, max_docs=20):
    """Read EPUB files by extracting text from XHTML/HTML chapters."""
    path = os.path.expanduser(path)
    title = os.path.basename(path)
    parts = [f"# EPUB文件: {title}"]
    with zipfile.ZipFile(path, "r") as zf:
        names = [name for name in zf.namelist() if name.lower().endswith((".xhtml", ".html", ".htm"))]
        chapter_names = [name for name in names if "/text/" in name.lower() or "/chapter" in name.lower() or "/chap" in name.lower()]
        selected = chapter_names or names
        used = 0
        for name in selected:
            if used >= max_docs:
                break
            try:
                raw = zf.read(name).decode("utf-8", errors="replace")
            except Exception:
                try:
                    raw = zf.read(name).decode("utf-8-sig", errors="replace")
                except Exception:
                    continue
            text = _strip_html_text(raw)
            if not text or len(text) < 20:
                continue
            heading = os.path.splitext(os.path.basename(name))[0]
            parts.append(f"\n## {heading}\n\n{text[:8000]}")
            used += 1
        if used == 0:
            raise ValueError("没有在 EPUB 里解析到可用正文")
        if len(selected) > used:
            parts.append(f"\n> 只提取前 {used} 个正文文件用于预览。")
    return "\n".join(parts), title


def detect_source_type(raw):
    """Detect whether input is a URL, file path, or plain text."""
    raw = raw.strip()
    # Check if it's a directory
    expanded = os.path.expanduser(raw)
    if os.path.isdir(expanded):
        return 'directory'
    # Check if it's a file path
    if os.path.isfile(expanded):
        ext = os.path.splitext(expanded)[1].lower()
        return 'file'
    # Check if it's a URL
    for pattern, source_type in URL_PATTERNS:
        if pattern.match(raw):
            return 'url' if source_type == 'url' else raw.split('/')[2].split('.')[0]
    # Check if it's an Obsidian reference
    if raw.startswith('obsidian://') or raw.startswith('@obsidian'):
        return 'obsidian'
    return 'paste'


def _cell_to_text(value):
    """Convert spreadsheet cell values into compact plain text."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M").rstrip(" 00:00")
    return str(value).replace("\n", " ").strip()


def _markdown_table(rows):
    rows = [[_cell_to_text(cell) for cell in row] for row in rows]
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        return "(空表)"

    max_cols = min(max(len(row) for row in rows), 12)
    normalized = [(row + [""] * max_cols)[:max_cols] for row in rows]
    header = normalized[0]
    if not any(header):
        header = [f"列{i + 1}" for i in range(max_cols)]
        body = normalized
    else:
        body = normalized[1:]

    def esc(cell):
        return cell.replace("|", "\\|")

    lines = [
        "| " + " | ".join(esc(cell) for cell in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    lines.extend("| " + " | ".join(esc(cell) for cell in row) + " |" for row in body)
    return "\n".join(lines)


def read_spreadsheet(path, max_rows=120):
    """Read Excel/CSV-like spreadsheet files into Markdown tables."""
    path = os.path.expanduser(path)
    title = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            rows = list(csv.reader(f))[:max_rows + 1]
        note = "\n\n> 只预览前 %d 行。" % max_rows if len(rows) > max_rows else ""
        return f"# 表格文件: {title}\n\n{_markdown_table(rows[:max_rows])}{note}", title

    if ext in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        parts = [f"# 表格文件: {title}"]
        for sheet in workbook.worksheets:
            rows = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx > max_rows:
                    break
                rows.append(list(row))
            parts.append(f"\n## 工作表: {sheet.title}\n\n{_markdown_table(rows)}")
            if sheet.max_row and sheet.max_row > max_rows:
                parts.append(f"\n> 该工作表共约 {sheet.max_row} 行，只预览前 {max_rows} 行。")
        workbook.close()
        return "\n".join(parts), title

    if ext == ".xls":
        import xlrd
        workbook = xlrd.open_workbook(path)
        parts = [f"# 表格文件: {title}"]
        for sheet in workbook.sheets():
            rows = [sheet.row_values(i) for i in range(min(sheet.nrows, max_rows))]
            parts.append(f"\n## 工作表: {sheet.name}\n\n{_markdown_table(rows)}")
            if sheet.nrows > max_rows:
                parts.append(f"\n> 该工作表共 {sheet.nrows} 行，只预览前 {max_rows} 行。")
        return "\n".join(parts), title

    raise ValueError(f"不支持的表格格式: {ext}")


def read_file(path):
    """Read a single file, return (content, title) or raises."""
    path = os.path.expanduser(path)
    ext = os.path.splitext(path)[1].lower()
    if ext in SPREADSHEET_EXTENSIONS or ext == ".csv":
        return read_spreadsheet(path)
    if ext == ".epub":
        return read_epub(path)
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    title = os.path.basename(path)
    return content, title


def read_directory(dir_path):
    """Recursively read all readable files in a directory.

    Returns (combined_content, dir_name).
    """
    dir_path = os.path.expanduser(dir_path)
    dir_name = os.path.basename(dir_path)
    parts = [f"# 目录: {dir_name}\n"]

    # Walk directory, collect files
    file_list = []
    for root, dirs, files in os.walk(dir_path):
        # Skip hidden directories and common non-source dirs
        dirs[:] = [d for d in dirs if not d.startswith('.') and d not in (
            '__pycache__', 'node_modules', 'venv', '.venv', 'env', '.env',
            'dist', 'build', '.git', '.svn', 'target', 'bin', 'obj',
            'vendor', '.mypy_cache', '.pytest_cache', '.ruff_cache',
            'egg-info', '.egg-info',
        )]
        for f in sorted(files):
            ext = os.path.splitext(f)[1].lower()
            if (ext in READABLE_EXTENSIONS or ext in SPREADSHEET_EXTENSIONS) and not f.startswith('.'):
                file_list.append(os.path.join(root, f))

    for fpath in file_list:
        rel_path = os.path.relpath(fpath, dir_path)
        try:
            ext = os.path.splitext(fpath)[1].lower()
            if ext in SPREADSHEET_EXTENSIONS or ext == ".csv":
                content, _ = read_spreadsheet(fpath)
            elif ext == ".epub":
                content, _ = read_epub(fpath)
            else:
                with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
            parts.append(f"\n## {rel_path}\n\n```\n{content}\n```\n")
        except Exception as e:
            parts.append(f"\n## {rel_path}\n\n[读取失败: {e}]\n")

    if len(parts) == 1:
        parts.append("(目录为空或无可用文件)")

    return "\n".join(parts), dir_name


def read_input(raw):
    """Read and return (content, source_type, title) from any input."""
    raw = raw.strip()
    source_type = detect_source_type(raw)

    if source_type == 'directory':
        content, title = read_directory(raw)
        return content, 'directory', title

    if source_type == 'file':
        content, title = read_file(raw)
        return content, 'file', title

    # For URLs and paste, return the raw content directly
    return raw, source_type, ''
